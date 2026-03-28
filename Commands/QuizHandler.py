import re
import os
import asyncio
import zipfile
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update
from telegram.ext import ContextTypes
from telegram.error import BadRequest, Forbidden, TimedOut

# ── Config ─────────────────────────────────────────────────────────────────────
QUIZ_GROUP_ID           = -1003818345268      # ← replace with your QuizGroup ID
BOT_MANAGEMENT_GROUP_ID = -1002359766306

MASTER_EXCEL_PATH = "UserScore/quiz_master.xlsx"         # srno | filename | numberofquestion | language
QUIZ_FOLDER       = "quiz_files"               # folder where individual quiz .xlsx files live

os.makedirs(QUIZ_FOLDER, exist_ok=True)

# ── In-memory session ─────────────────────────────────────────────────────────
quiz_session: dict = {"active": False}


# ══════════════════════════════════════════════════════════════════════════════
#  STYLING HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row_num: int, col_count: int):
    fill   = PatternFill("solid", start_color="1F4E79")
    font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    border = _thin_border()
    for col in range(1, col_count + 1):
        cell           = ws.cell(row_num, col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border


# ══════════════════════════════════════════════════════════════════════════════
#  MASTER EXCEL  —  single source of truth for all quiz file metadata
# ══════════════════════════════════════════════════════════════════════════════

MASTER_HEADERS = ["srno", "filename", "numberofquestion", "language"]


def _ensure_master_excel():
    if not os.path.exists(MASTER_EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QuizFiles"
        ws.append(MASTER_HEADERS)
        _style_header_row(ws, 1, len(MASTER_HEADERS))
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12
        wb.save(MASTER_EXCEL_PATH)


def _read_master_rows() -> list[dict]:
    """Return all data rows as list of dicts, sorted by srno."""
    _ensure_master_excel()
    wb   = openpyxl.load_workbook(MASTER_EXCEL_PATH)
    ws   = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append({
                "srno"            : row[0],
                "filename"        : row[1] or "",
                "numberofquestion": row[2] or 0,
                "language"        : row[3] or "",
            })
    rows.sort(key=lambda r: (r["srno"] is None, r["srno"]))
    return rows


def _rebuild_master_from_disk() -> list[dict]:
    """
    Scan QUIZ_FOLDER for every .xlsx, count their actual rows,
    then rewrite master Excel with clean sequential srno (sorted by filename).
    Preserves existing language info.  Returns the rebuilt rows.
    """
    _ensure_master_excel()

    # Keep existing language mapping so we don't lose it
    existing_lang = {r["filename"]: r["language"] for r in _read_master_rows()}

    files = sorted(
        f[:-5] for f in os.listdir(QUIZ_FOLDER) if f.endswith(".xlsx")
    )

    rows = []
    for i, base in enumerate(files, start=1):
        rows.append({
            "srno"            : i,
            "filename"        : base,
            "numberofquestion": _count_questions_in_excel(base),
            "language"        : existing_lang.get(base, "En"),
        })

    _write_master_rows(rows)
    return rows


def _write_master_rows(rows: list[dict]):
    """Completely rewrite the master Excel from a list of row dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QuizFiles"
    ws.append(MASTER_HEADERS)
    _style_header_row(ws, 1, len(MASTER_HEADERS))

    border   = _thin_border()
    alt_fill = PatternFill("solid", start_color="DCE6F1")

    for idx, r in enumerate(rows):
        ws.append([r["srno"], r["filename"], r["numberofquestion"], r["language"]])
        row_i = ws.max_row
        fill  = alt_fill if idx % 2 == 1 else None
        for col in range(1, 5):
            cell           = ws.cell(row_i, col)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 3, 4) else "left",
                vertical="center"
            )
            cell.border = border
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    wb.save(MASTER_EXCEL_PATH)


def _update_master_for_file(filename: str, num_questions: int, language: str):
    """
    Update or insert one file entry, then reassign all srno values
    sequentially (sorted by filename a-z) and rewrite master Excel.
    """
    _ensure_master_excel()

    # Load all rows
    rows = _read_master_rows()

    # Update existing or append new
    found = False
    for r in rows:
        if r["filename"] == filename:
            r["numberofquestion"] = num_questions
            r["language"]         = language
            found = True
            break
    if not found:
        rows.append({
            "srno"            : None,
            "filename"        : filename,
            "numberofquestion": num_questions,
            "language"        : language,
        })

    # Sort by filename and reassign srno 1-based
    rows.sort(key=lambda r: r["filename"].lower())
    for i, r in enumerate(rows, start=1):
        r["srno"] = i

    _write_master_rows(rows)


# ══════════════════════════════════════════════════════════════════════════════
#  QUIZ EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

QUIZ_HEADERS = ["srno", "question", "option1", "option2", "option3", "option4", "answer"]


def _save_quiz_excel(filename: str, questions: list[dict]) -> str:
    """Append questions to (or create) a quiz Excel file. Returns filepath."""
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(QUIZ_FOLDER, filename)

    if os.path.exists(filepath):
        wb       = openpyxl.load_workbook(filepath)
        ws       = wb.active
        start_no = ws.max_row   # header=row1, so next question srno = current max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        ws.append(QUIZ_HEADERS)
        _style_header_row(ws, 1, len(QUIZ_HEADERS))
        start_no = 1

    border   = _thin_border()
    alt_fill = PatternFill("solid", start_color="EBF1DE")

    for i, q in enumerate(questions, start=start_no):
        ws.append([
            i,
            q["question"],
            q["option1"],
            q["option2"],
            q.get("option3", ""),
            q.get("option4", ""),
            q["answer"],
        ])
        row_idx = ws.max_row
        fill    = alt_fill if (i % 2 == 0) else None
        for col in range(1, len(QUIZ_HEADERS) + 1):
            cell           = ws.cell(row_idx, col)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = border
            if fill:
                cell.fill = fill

    col_widths = [8, 60, 30, 30, 30, 30, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(filepath)
    return filepath


def _count_questions_in_excel(filename: str) -> int:
    """Count actual data rows (excluding header) in a quiz Excel file."""
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(QUIZ_FOLDER, filename)
    if not os.path.exists(filepath):
        return 0
    wb    = openpyxl.load_workbook(filepath, read_only=True)
    ws    = wb.active
    count = sum(
        1 for row in ws.iter_rows(min_row=2, values_only=True)
        if any(c is not None for c in row)
    )
    wb.close()
    return count


# ══════════════════════════════════════════════════════════════════════════════
#  ZIP HELPER  —  bundles all quiz files + master into one zip
# ══════════════════════════════════════════════════════════════════════════════

def _build_zip() -> str | None:
    """
    Zip every .xlsx in QUIZ_FOLDER plus quiz_master.xlsx.
    Returns path to a temp zip file, or None if nothing exists.
    """
    quiz_files = [
        os.path.join(QUIZ_FOLDER, f)
        for f in os.listdir(QUIZ_FOLDER) if f.endswith(".xlsx")
    ]

    if not quiz_files and not os.path.exists(MASTER_EXCEL_PATH):
        return None

    stamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_path = os.path.join(tempfile.gettempdir(), f"quiz_backup_{stamp}.zip")

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in sorted(quiz_files):
            zf.write(fp, arcname=os.path.join("quiz_files", os.path.basename(fp)))
        if os.path.exists(MASTER_EXCEL_PATH):
            zf.write(MASTER_EXCEL_PATH, arcname="quiz_master.xlsx")

    return zip_path


async def _send_zip(context, chat_id: int, caption: str):
    """Build ZIP and send it, then delete the temp file."""
    try:
        zip_path = _build_zip()
        if not zip_path:
            await context.bot.send_message(
                chat_id=chat_id,
                text="📂 No quiz files found on disk yet."
            )
            return
        with open(zip_path, "rb") as f:
            await context.bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=os.path.basename(zip_path),
                caption=caption,
            )
    except Exception as e:
        print(f"[_send_zip] {e}")
    finally:
        try:
            if zip_path and os.path.exists(zip_path):
                os.remove(zip_path)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
#  POLL PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _extract_poll_data(poll) -> dict | None:
    if poll is None:
        return None
    question = poll.question or ""
    options  = [o.text for o in (poll.options or [])]
    if len(options) < 2:
        return None
    correct_idx = getattr(poll, "correct_option_id", None)
    answer = (
        options[correct_idx]
        if correct_idx is not None and correct_idx < len(options)
        else options[0]
    )
    return {
        "question": question,
        "option1" : options[0] if len(options) > 0 else None,
        "option2" : options[1] if len(options) > 1 else None,
        "option3" : options[2] if len(options) > 2 else None,
        "option4" : options[3] if len(options) > 3 else None,
        "answer"  : answer,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  /addquiz  — sync master from disk → show list → send ZIP in background
# ══════════════════════════════════════════════════════════════════════════════

async def addquiz_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id != QUIZ_GROUP_ID:
        try:
            await update.message.reply_text("❌ This command is only for the Quiz Group.")
        except Exception:
            pass
        return

    if quiz_session.get("active"):
        try:
            await update.message.reply_text(
                "⚠️ A quiz session is already active!\n"
                f"File: <b>{quiz_session['filename']}</b>\n"
                f"Questions collected: <b>{len(quiz_session['questions'])}</b>\n\n"
                "Use /stop to finish first.",
                parse_mode="HTML"
            )
        except Exception:
            pass
        return

    # Always resync from disk so srno is never stale
    rows = _rebuild_master_from_disk()

    if rows:
        lines = []
        for r in rows:
            flag = "🇬🇧" if str(r["language"]).lower() in ("en", "english") else "🇮🇳"
            lines.append(
                f"<b>{r['srno']}.</b>  {r['filename']}  |  "
                f"{r['numberofquestion']} Qs  |  {flag} {r['language']}"
            )
        msg = (
            "📂 <b>Quiz Files</b>\n\n"
            + "\n".join(lines)
            + "\n\n━━━━━━━━━━━━━━━━━━━━\n"
            "Reply with a <b>number</b> to add questions to that file,\n"
            "or type <b>new</b> to create a new file."
        )
    else:
        msg = (
            "📂 <b>No quiz files yet.</b>\n\n"
            "Type <b>new</b> to create your first quiz file."
        )

    context.user_data["awaiting_quiz_selection"] = True

    try:
        await update.message.reply_text(msg, parse_mode="HTML")
    except Exception as e:
        print(f"[addquiz_command] {e}")

    # Send ZIP in background — doesn't block the reply above
    asyncio.create_task(
        _send_zip(
            context,
            chat_id,
            caption="📦 Current quiz backup (all files + master)"
        )
    )


# ══════════════════════════════════════════════════════════════════════════════
#  Text handler  —  selection → filename → language (multi-step)
# ══════════════════════════════════════════════════════════════════════════════

async def handle_quiz_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    print("uyt")
    if chat_id != QUIZ_GROUP_ID:
        return

    text = (update.message.text or "").strip()
    ud   = context.user_data

    # ── Step 1 ────────────────────────────────────────────────────────────
    if ud.get("awaiting_quiz_selection"):
        rows = _read_master_rows()

        if text.lower() == "new":
            ud["awaiting_quiz_selection"] = False
            ud["awaiting_new_filename"]   = True
            try:
                await update.message.reply_text(
                    "✏️ Enter a name for the new quiz file (without .xlsx):"
                )
            except Exception:
                pass
            return

        if text.isdigit():
            srno  = int(text)
            match = next((r for r in rows if r["srno"] == srno), None)
            if not match:
                try:
                    await update.message.reply_text(
                        f"❌ No file with Sr.No <b>{srno}</b>. Try again or type <b>new</b>.",
                        parse_mode="HTML"
                    )
                except Exception:
                    pass
                return

            ud["awaiting_quiz_selection"] = False
            _start_session(match["filename"], match["language"], match["srno"])
            try:
                await update.message.reply_text(
                    f"✅ Session started for <b>{match['filename']}</b> ({match['language']})\n\n"
                    "📩 Forward polls here — I'll collect them.\n"
                    "Send /stop when done.",
                    parse_mode="HTML"
                )
            except Exception:
                pass
            return

        try:
            await update.message.reply_text(
                "❓ Send a number or type <b>new</b>.", parse_mode="HTML"
            )
        except Exception:
            pass
        return

    # ── Step 2 ────────────────────────────────────────────────────────────
    if ud.get("awaiting_new_filename"):
        filename = re.sub(r'[\\/*?:"<>|]', "_", text).strip()
        if not filename:
            try:
                await update.message.reply_text("❌ Invalid filename. Try again:")
            except Exception:
                pass
            return

        ud["awaiting_new_filename"] = False
        ud["pending_filename"]      = filename
        ud["awaiting_language"]     = True
        try:
            await update.message.reply_text(
                f"🌐 Language for <b>{filename}</b>?\n\n"
                "Reply:\n• <b>En</b> for English\n• <b>Hi</b> for Hindi",
                parse_mode="HTML"
            )
        except Exception:
            pass
        return

    # ── Step 3 ────────────────────────────────────────────────────────────
    if ud.get("awaiting_language"):
        lang_raw = text.strip().lower()
        lang = "En" if lang_raw in ("en", "english") else \
               "Hi" if lang_raw in ("hi", "hindi")   else None

        if lang is None:
            try:
                await update.message.reply_text(
                    "❓ Please reply <b>En</b> or <b>Hi</b>.", parse_mode="HTML"
                )
            except Exception:
                pass
            return

        filename = ud.pop("pending_filename")
        ud.pop("awaiting_language", None)

        # srno will be assigned automatically when master is rebuilt on /stop
        _start_session(filename, lang, srno=None)

        try:
            await update.message.reply_text(
                f"✅ New file <b>{filename}.xlsx</b> ({lang}) ready.\n\n"
                "📩 Forward polls here — I'll collect them.\n"
                "Send /stop when done.",
                parse_mode="HTML"
            )
        except Exception:
            pass
        return


def _start_session(filename: str, language: str, srno):
    quiz_session["active"]    = True
    quiz_session["filename"]  = filename
    quiz_session["language"]  = language
    quiz_session["srno"]      = srno
    quiz_session["questions"] = []


# ══════════════════════════════════════════════════════════════════════════════
#  Poll collector
# ══════════════════════════════════════════════════════════════════════════════
async def handle_quiz_poll(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id != QUIZ_GROUP_ID:
        return
    if not quiz_session.get("active"):
        return

    poll = update.message.poll if update.message else None
    data = _extract_poll_data(poll)

    if data is None:
        try:
            await update.message.reply_text("⚠️ Could not read this poll. Skipping.")
        except Exception:
            pass
        return

    quiz_session["questions"].append(data)
    count = len(quiz_session["questions"])

    # Build a compact 4-option display (show None as –)
    opts = "\n".join(
        f"  {'ABCD'[i]}. {data[f'option{i+1}'] or '–'}"
        for i in range(4)
    )

    try:
        await update.message.reply_text(
            f"✅ <b>Q{count} saved</b>\n\n"
            f"<b>Q:</b> {data['question']}\n\n"
            f"{opts}\n\n"
            f"<b>✔ Answer:</b> {data['answer']}",
            parse_mode="HTML"
        )
    except Exception:
        pass
# ══════════════════════════════════════════════════════════════════════════════
#  /stop  — save file → update master (correct srno) → send master + ZIP
# ══════════════════════════════════════════════════════════════════════════════

async def stop_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id != QUIZ_GROUP_ID:
        try:
            await update.message.reply_text("❌ Not allowed here.")
        except Exception:
            pass
        return

    if not quiz_session.get("active"):
        try:
            await update.message.reply_text("ℹ️ No active quiz session.")
        except Exception:
            pass
        return

    filename  = quiz_session["filename"]
    language  = quiz_session["language"]
    questions = quiz_session["questions"]

    if not questions:
        quiz_session["active"] = False
        try:
            await update.message.reply_text(
                "⚠️ Session ended — no questions were collected."
            )
        except Exception:
            pass
        return

    try:
        status_msg = await update.message.reply_text("⏳ Saving quiz…")
    except Exception:
        status_msg = None

    try:
        # 1. Append questions to quiz Excel
        _save_quiz_excel(filename, questions)

        # 2. Count actual total on disk
        total_questions = _count_questions_in_excel(filename)

        # 3. Update master: insert/update this file, then fix ALL srno values
        _update_master_for_file(filename, total_questions, language)

        # 4. Read back what srno was assigned to this file
        updated_rows  = _read_master_rows()
        file_row      = next((r for r in updated_rows if r["filename"] == filename), None)
        assigned_srno = file_row["srno"] if file_row else "?"

        # 5. Reset session
        quiz_session["active"] = False

        result = (
            f"✅ <b>Quiz Saved!</b>\n\n"
            f"├ Sr.No     → <b>{assigned_srno}</b>\n"
            f"├ File      → <b>{filename}.xlsx</b>\n"
            f"├ Language  → <b>{language}</b>\n"
            f"├ New Qs    → <b>{len(questions)}</b>\n"
            f"└ Total Qs  → <b>{total_questions}</b>\n\n"
            f"📊 Sending master + full ZIP backup…"
        )

        if status_msg:
            try:
                await status_msg.edit_text(result, parse_mode="HTML")
            except Exception:
                pass

        # 6. Send updated master Excel immediately
        with open(MASTER_EXCEL_PATH, "rb") as f:
            await context.bot.send_document(
                chat_id=chat_id,
                document=f,
                filename="quiz_master.xlsx",
                caption=(
                    f"📊 Master file — {len(updated_rows)} quiz file(s) · "
                    f"srno reassigned & corrected"
                )
            )

        # 7. Send full ZIP in background (non-blocking)
        asyncio.create_task(
            _send_zip(
                context,
                chat_id,
                caption=(
                    f"📦 Full backup — {len(updated_rows)} file(s)  |  "
                    f"Latest: {filename} (+{len(questions)} Qs, total {total_questions})"
                )
            )
        )

    except Exception as e:
        print(f"[stop_command] Error: {e}")
        quiz_session["active"] = False
        try:
            msg = f"❌ Error saving: {e}"
            if status_msg:
                await status_msg.edit_text(msg)
            else:
                await update.message.reply_text(msg)
        except Exception:
            pass

