from telegram import Update, PollAnswer, Poll, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import  Application,filters, CommandHandler,MessageHandler, PollAnswerHandler, CallbackQueryHandler, ContextTypes
from telegram.error import BadRequest, Forbidden, TimedOut
from Handlers.config import  DIFFICULTY_MAP,ALLOWED_FILES,Reasoning_Kb0,Upsc_keyboard2,Upsc_keyboard0,Upsc_keyboard1, StartingSubject0,StartingSubject1, Nda_keyboard0, Nda_keyboard1, Nda_keyboard2, Topic_Kb0, Topic_Kb1, Topic_Kb2
import os
import json
import pandas as pd
import openpyxl
from typing import Final
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import asyncio,re,random,time


TOKEN: Final = '7938454369:AAHvTD7J-C2OozXpu4XQc-rvjQNOLhgrO6s'
#TOKEN: Final = '7007935023:AAENkGaklw6LMJA_sfhVZhnoAgIjW4lDTBc'
BOT_USERNAME: Final = '@slizzyy_bot'

ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690,-1001817635995]
# EXCEL_FILE = 'Data/SYNO5.xlsx'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'Data', 'SYNO5.xls')
MNTH_SCORE_FILE="UserScore/MonthlyUserScore.xlsx"
SCORE_FILE="UserScore/user_scores.xlsx"
groupsendid = -1002114430690
botManagementGroupId =  -1002359766306
new_timer = 20
Promotion = False
final_poll_responses = {}

GROUPS_FILE = "groups.json"
group_data_file = "group_data.json"

if os.path.exists(GROUPS_FILE):
    with open(GROUPS_FILE, "r") as f:
        registered_groups = set(json.load(f))
else:
    registered_groups = set()


if os.path.exists(group_data_file):
    with open(group_data_file, "r") as f:
        group_data = json.load(f)
else:
    group_data = {}


def save_groups():
    with open(GROUPS_FILE, "w") as f:

        json.dump(list(registered_groups), f)


def save_group_data():
    with open(group_data_file, "w") as f:
        json.dump(group_data, f)


newuploadedexcelfile = "Data/SYNO5.xlsx"
quiz_state = {}
correct_users = {}  
quiz_scores = {} 
selected_poll_count = 10
selected_quizscore_count=0
active_poll=1
answers_received = defaultdict(int) 
is_quiz_active = False  
chat_id = None  
selected_time_limit = 10  
unanswered_poll = 0
cancel_active = False
display_chat=0
Quiz_grammar_type =''
quiz_kick= False
quiz_tasks = {}
StudyStuffgrp=False
commandfunctionpass = 1
isNewQuizStarted =1
used_srnos = set()
difficulty_message = "subject"


def get_chat_state(chat_id):
    if chat_id not in quiz_state:
        quiz_state[chat_id] = {
            "is_active": False,
            "selected_polls": [],
            "correct_users": {},
        }
    return quiz_state[chat_id]


def reset_used_srnos():
    global used_srnos
    used_srnos.clear()

def escape_markdown(text: str) -> str:
    return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!])', r'\\\1', text)

def load_scores():
    if not os.path.exists(SCORE_FILE):
        return []

    workbook = openpyxl.load_workbook(SCORE_FILE)
    sheet = workbook.active

    scores = []
    for row in range(2, sheet.max_row + 1):  

        user_id = sheet.cell(row=row, column=2).value
        username = sheet.cell(row=row, column=3).value
        score = sheet.cell(row=row, column=4).value
        round = sheet.cell(row=row, column=5).value
        if user_id and username and score and round is not None:
            scores.append((user_id, username, score,round))

    workbook.close()
    return scores

def load_quiz_data(file_path, selected_poll_count):
    global used_srnos
    try:
        df = pd.read_excel(file_path)
        
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        unique_rows = df[~df['srno'].isin(used_srnos)]
        if len(unique_rows) < selected_poll_count:
            print("Not enough unique rows available.")
            selected_poll_count = len(unique_rows)
        
        selected_rows = unique_rows.sample(n=selected_poll_count)
        used_srnos.update(selected_rows['srno'].tolist())
        
        polls = []
        for _, row in selected_rows.iterrows():
            options = [row["option1"], row["option2"], row["option3"], row["option4"]]
            random.shuffle(options) 
            poll = {
                "question": row["question"],
                "options": options,
                "correct_answer": row["answer"],
                "meaning": row.get("meaning", "nan") 
            }
            polls.append(poll)
            
        return polls
    except Exception as e:
        print(e)

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    await context.bot.send_message(chat_id, text="use /startquiz Command To start the Quizzes")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    await context.bot.send_message(chat_id, text="Drop a Message on @O000000000O00000000O This id")

async def start_game_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global is_quiz_active,quiz_state, correct_users,commandfunctionpass, chat_id, unanswered_poll,cancel_active,selected_quizscore_count,quiz_kick
        cancel_active = False
        quiz_kick= False
        commandfunctionpass = 1
        reset_used_srnos()
        chat_id = update.message.chat.id

        if chat_id in quiz_state:
            await update.message.chat.send_message('A quiz is already running in this group. Wait or Cancel it by /cancelquiz')
            return
        if chat_id not in quiz_state:
            quiz_state[chat_id] = {
                "is_active": True,
                "active":True,
                "polls": [],
                "scores": {},  
                "quiz_kick": False,  
                "cancel_active": False,  
                "consecutive_unanswered": 0,
                "selected_poll_count": 0,  
            }

        selected_quizscore_count=0
        correct_users.clear() 
        reply_markup =  InlineKeyboardMarkup(StartingSubject0())
        try:
            await update.message.chat.send_message('Select the Quiz type:', reply_markup=reply_markup)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)

async def handle_allsizzlescore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    if chat_id == groupsendid and update.message.reply_to_message and update.message.reply_to_message.document:
        try:
            file = await context.bot.get_file(update.message.reply_to_message.document.file_id)
            downloaded_path = await file.download_to_drive()

            # Replace existing userscore.xlsx
            if os.path.exists(SCORE_FILE):
                os.remove(SCORE_FILE)

            os.rename(downloaded_path, SCORE_FILE)

            await update.message.reply_text("✅ Score file updated successfully.")
        except Exception as e:
            await update.message.reply_text(f"❌ Failed to update score file:\n`{e}`", parse_mode="Markdown")
    else:
        await update.message.reply_text("⚠️ Please reply to an .xlsx file in the allowed group.")

async def handle_updatesizzlescore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    if chat_id == groupsendid and update.message.reply_to_message and update.message.reply_to_message.document:
        try:
            file = await context.bot.get_file(update.message.reply_to_message.document.file_id)
            downloaded_path = await file.download_to_drive()

            # Replace existing userscore.xlsx
            if os.path.exists(MNTH_SCORE_FILE):
                os.remove(MNTH_SCORE_FILE)

            os.rename(downloaded_path, MNTH_SCORE_FILE)

            await update.message.reply_text("✅ Score file updated successfully.")
        except Exception as e:
            await update.message.reply_text(f"❌ Failed to update score file:\n`{e}`", parse_mode="Markdown")
    else:
        await update.message.reply_text("⚠️ Please reply to an .xlsx file in the allowed group.")

async def handle_difficulty_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global EXCEL_FILE, Quiz_grammar_type,StudyStuffgrp,difficulty_message
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name
      

        await query.answer()
        difficulty_message = ''
        chat_id = query.message.chat.id
        if chat_id not in quiz_state:
            quiz_state[chat_id] = {
                "is_active": True,
                "active":True,
                "polls": [],
                "scores": {},  
                "quiz_kick": False,  
                "cancel_active": False,  
                "consecutive_unanswered": 0,
                "selected_poll_count": 0,  
            }
        data = query.data
        if data not in DIFFICULTY_MAP:
            await query.edit_message_text("⚠️ Invalid selection.")
            return

        EXCEL_FILE, difficulty_message = DIFFICULTY_MAP[data]
        Quiz_grammar_type = difficulty_message
        
        if(Quiz_grammar_type !='Reasoning' and Quiz_grammar_type !='Maths'):
            time_keyboard = [
            
            [InlineKeyboardButton("15 Seconds", callback_data='time_15')],
            [InlineKeyboardButton("20 Seconds", callback_data='time_20')],
            [InlineKeyboardButton("25 Seconds", callback_data='time_25')],
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            
            ]
        else:
            time_keyboard = [
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            [InlineKeyboardButton("45 Seconds", callback_data='time_45')],
            [InlineKeyboardButton("60 Seconds", callback_data='time_60')],
            [InlineKeyboardButton("90 Seconds", callback_data='time_90')],
            ]
        reply_markup = InlineKeyboardMarkup(time_keyboard)
        if not quiz_state[chat_id]["is_active"]:
            try:
                await query.edit_message_text(f'Quiz is Already Running in this Chat. Wait or Cancel it /cancelquiz')
            
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'Quiz is Already Running in this Chat. Wait or Cancel it /cancelquiz')
        else:
            try:
                await query.edit_message_text(f'@{username} Chooses the {difficulty_message}  \n\n Select the time limit for each poll:', reply_markup=reply_markup)
            
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Chooses the {difficulty_message}  \n\n Select the time limit for each poll:', reply_markup=reply_markup)
    
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)

async def handle_type_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        await query.answer()
        await query.answer()
        difficulty_message = ''

        if query.data == 'type_NDA0':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard0())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 1 \n Select the Grammar Quiz type:', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 1 \n Select the Grammar Quiz type:', reply_markup=reply_markup)
        

        elif query.data == 'type_NDA1':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard1())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
        

        elif query.data == 'type_NDA2':
            reply_markup = InlineKeyboardMarkup(Nda_keyboard2())
            try:
                await query.edit_message_text(f'@{username} selected NDA-CDS Phase 2 \n\n Select the Grammar Quiz type:', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected NDA-CDS Phase 2\n Select the Grammar Quiz type:', reply_markup=reply_markup)
        
        elif query.data == 'type_startsubj0':
            try:
                 await query.edit_message_text( "📘 Select the Quiz Subject:", reply_markup=InlineKeyboardMarkup(StartingSubject0()))
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message( "📘 Select the Quiz Subject:", reply_markup=InlineKeyboardMarkup(StartingSubject0()))
        

        elif query.data == 'type_startsubj1':
            try:
                 await query.edit_message_text( "📘 Select the Quiz Subject:", reply_markup=InlineKeyboardMarkup(StartingSubject1()))
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message( "📘 Select the Quiz Subject:", reply_markup=InlineKeyboardMarkup(StartingSubject0()))
        elif query.data == 'type_topic0':
            reply_markup = InlineKeyboardMarkup(Topic_Kb0())
            try:
                await query.edit_message_text(f'@{username} selected Topic Phase 1 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected Topic Phase 1 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
        

        elif query.data == 'type_topic1':
            reply_markup = InlineKeyboardMarkup(Topic_Kb1())
            try:
                await query.edit_message_text(f'@{username} selected Topic Phase 2 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected Topic Phase 2 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
        

        elif query.data == 'type_topic2':
            reply_markup = InlineKeyboardMarkup(Topic_Kb2())
            try:
                await query.edit_message_text(f'@{username} selected Topic Phase 3 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected Topic Phase 3 \n\n Select the Quiz Topic :', reply_markup=reply_markup)
        
        
        elif query.data == 'type_BASIC':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms')],
                [InlineKeyboardButton("Preposition", callback_data='difficulty_prepo')],
                [InlineKeyboardButton("Antonyms", callback_data='difficulty_antonyms')],
                [InlineKeyboardButton("Spelling Correction 2.0", callback_data='difficulty_spellcorr')],
                [InlineKeyboardButton("Daily Life Idioms", callback_data='difficulty_shortIdiom')],
                [InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected Basic Grammar \n Select the Grammar Quiz type:', reply_markup=reply_markup)   
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected Basic Grammar \n Select the Grammar Quiz type:', reply_markup=reply_markup)
        elif query.data == 'type_Upsc2':
            
            reply_markup = reply_markup=InlineKeyboardMarkup(Upsc_keyboard2())
            
            try:
                await query.edit_message_text(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
       
        elif query.data == 'type_Upsc1':
            
            reply_markup = reply_markup=InlineKeyboardMarkup(Upsc_keyboard1())
            
            try:
                await query.edit_message_text(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
        elif query.data == 'type_reasoning':
            
            reply_markup = reply_markup=InlineKeyboardMarkup(Reasoning_Kb0())
            
            try:
                await query.edit_message_text(f'@{username} Selected Reasoning \n Select the  Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
        
        elif query.data == 'type_Upsc0':
            
            reply_markup = reply_markup=InlineKeyboardMarkup(Upsc_keyboard0())
            
            try:
                await query.edit_message_text(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
        elif query.data == 'type_Cgl':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Reasoning", callback_data='difficulty_cglReasoning')],
                [InlineKeyboardButton("English", callback_data='difficulty_cglEnglish')],
                [InlineKeyboardButton("General Awareness", callback_data='difficulty_cglGk')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected SSC - CGL/CHSL \n Select the Grammar Quiz type', reply_markup=reply_markup)
                
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected UPSC \n Select the Grammar Quiz type', reply_markup=reply_markup)
       
        elif query.data == 'type_Neet':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Chemistry", callback_data='difficulty_neetchemistry')]
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected Jee and Neet \n Select the Quiz type:', reply_markup=reply_markup)   
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected  Jee and  Neet \n Select the Quiz type:', reply_markup=reply_markup)
        elif query.data == 'type_History':
            
            difficulty_keyboard = [
                [InlineKeyboardButton("Ancient History", callback_data='difficulty_historyAncient')],
                [InlineKeyboardButton("Medieval History", callback_data='difficulty_historyMedieval')],
                [InlineKeyboardButton("Modern History", callback_data='difficulty_historyModern')]
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.edit_message_text(f'@{username} Selected History \n Select the History Period type:', reply_markup=reply_markup)   
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} Selected History \n Select the History Period type:', reply_markup=reply_markup)
        
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
          
async def handle_time_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global selected_time_limit
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name
        chat_id = query.message.chat.id
        if chat_id not in quiz_state:
            quiz_state[chat_id] = {
                "is_active": True,
                "active":True,
                "polls": [],
                "scores": {},  
                "quiz_kick": False,  
                "cancel_active": False,  
                "consecutive_unanswered": 0,
                "selected_poll_count": 0,  
            }
        await query.answer()
        time_mapping = {
            
            'time_15': 15,
            'time_20': 20,
            'time_25': 25,
            'time_30': 30,
            'time_45': 45,
            'time_60': 60,
            'time_90': 90,
        }
        selected_time_limit = time_mapping.get(query.data, 10)
        if not quiz_state[chat_id]["is_active"]:
            return
        else:
            quiz_state[chat_id]["selectedtime"] = selected_time_limit
        keyboard = [
            [InlineKeyboardButton("15 Words", callback_data='15')],
            [InlineKeyboardButton("25 Words", callback_data='25')],
            [InlineKeyboardButton("35 Words", callback_data='35')],
            [InlineKeyboardButton("50 Words", callback_data='50')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        if not quiz_state[chat_id]["is_active"]:
            try:
                await query.edit_message_text(f'Quiz is Already Running in this Chat. Wait or Cancel it /cancelquiz')
            
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'Quiz is Already Running in this Chat. Wait or Cancel it /cancelquiz')
        else:
            try:
                await query.edit_message_text(f'@{username} selected {selected_time_limit} second To complete one quiz. \n\n How many rounds?', reply_markup=reply_markup)
            
            except (BadRequest, Forbidden, TimedOut) as e:
                await query.message.chat.send_message(f'@{username} selected {selected_time_limit} second To complete one quiz. \n\n How many rounds?', reply_markup=reply_markup)
       

    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)

async def handle_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global quiz_state,quiz_tasks
        query = update.callback_query
        chat_id = query.message.chat_id  # Identify which chat is running the quiz
        username = query.from_user.username or query.from_user.first_name
        if chat_id not in quiz_state:
            print("enter here 1")
            quiz_state[chat_id] = {"active": False}

        if not quiz_state[chat_id]["active"]:
            print("enter here 2")
            await query.answer("Please start a new quiz with /startquiz")
            return

        selected_poll_count = int(query.data)
        if not quiz_state[chat_id]["is_active"]:
            return
        quiz_state[chat_id]["total_rounds"] =  selected_poll_count
        quiz_state[chat_id]["active"] = True
        quiz_state[chat_id]["polls"] = []
 
        message = (
    f"🚀 @{username} has started the {difficulty_message} quiz!\n\n"
    f"🕒 Time per question: {selected_time_limit} seconds\n"
    f"🔢 Total rounds: {selected_poll_count}\n\n"
    f"Get ready — the quiz begins now! 🎯"
)
        
        await query.edit_message_text(text=message)
        selected_polls = load_quiz_data(EXCEL_FILE, selected_poll_count)
        if not quiz_state[chat_id]["is_active"]:
            return
        task = asyncio.create_task(send_quiz_polls(chat_id, selected_polls, context))
        quiz_tasks[chat_id] = task

    except Exception as e:
        print(f"Error in handle_button_click: {e}")

async def handle_New_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global newuploadedexcelfile,isNewQuizStarted,quiz_tasks,quiz_state,new_timer
        query = update.callback_query
        chat_id = query.message.chat.id
        if chat_id in quiz_state and quiz_state[chat_id]["active"]:
            quiz_state[chat_id]["active"] = False
            quiz_state.pop(chat_id, None)
            
        username = query.from_user.username or query.from_user.first_name
        selected_poll_count = 15
        selected_time_limit = int(new_timer)  # ✅ Make sure THIS line is present
      
        EXCEL_FILE = newuploadedexcelfile
        difficulty_message = "NEW"

        quiz_state[chat_id] = {

            "total_rounds": selected_poll_count,
            "selectedtime": selected_time_limit,
            "active": True,
            "polls": [],
            "is_active": True
        }
        
        message = (
            f"🚀 @{username} has started the {difficulty_message} quiz!\n\n"
            f"🕒 Time per question: {selected_time_limit} seconds\n"
            f"🔢 Total rounds: {selected_poll_count}\n\n"
            f"Get ready — the quiz begins now! 🎯"
        )

        await query.edit_message_text(text=message)
        selected_polls = load_quiz_data(EXCEL_FILE, selected_poll_count)
        if not quiz_state[chat_id]["is_active"]:
            return
        if chat_id in quiz_tasks:
            task = quiz_tasks[chat_id]
            if not task.done():
                task.cancel()
        print(f"Cancelled previous quiz task for chat_id {chat_id}")
        task  = asyncio.create_task(send_quiz_polls(chat_id, selected_polls, context))
        quiz_tasks[chat_id] = task

    except Exception as e:
        print(f"Error in handle_New_button_click: {e}")

async def send_quiz_polls(chat_id, polls, context):
    try:
        if not quiz_state[chat_id]["is_active"]:
            await context.bot.send_message(chat_id, text="Quiz is already running in this group. Please wait or cancel with /cancelquiz")
            return
        quiz_state[chat_id].setdefault("polls", [])
        quiz_scores[chat_id] = {}
        for i, poll in enumerate(polls):
            quiz_state[chat_id]["is_active"] = False
            if not quiz_state[chat_id]["active"]:
                print(f"Quiz canceled in {chat_id}, stopping polls.")
                break

            try:
                # Ensure all fields are stringified
                question = stringify(poll.get('question'))
                options = [stringify(opt) for opt in poll.get('options', [])]
                correct_answer = stringify(poll.get('correct_answer'))
                meaning = stringify(poll.get('meaning'))

                display_question = question
                if len(question) > 253:
                    await context.bot.send_message(chat_id, text=question)
                    display_question = "Choose Correct Answer"

                # Attempt to send poll
                poll_message = await context.bot.send_poll(
                    chat_id=chat_id,
                    question=f"{i+1}/{len(polls)}: {display_question}",
                    options=options,
                    is_anonymous=False,
                    allows_multiple_answers=False,
                    type=Poll.QUIZ,
                    correct_option_id=options.index(correct_answer)
                )

                quiz_state[chat_id]["polls"].append({
                    "poll_id": poll_message.poll.id,
                    "question": question,
                    "correct_answer": correct_answer,
                    "options": options,
                    "meaning": meaning,
                    "responses": {},
                    "poll_message": poll_message
                })
                if(i==7 and Promotion):
                    await context.bot.send_message(chat_id, text="\nGet All Quiz With topics: https://t.me/+BIGWj3dm7vA1NTNl")
                
                await countdown_and_close_poll(chat_id, poll_message, context)
                await asyncio.sleep(2)

            except BadRequest as e:
                await context.bot.send_message(chat_id, text="❗ I don't have permission to create polls in this group.\nPlease make me admin or allow 'Create Polls' permission.")
                break
            except Exception as e:
                print(f"Error sending poll: {e}")

        if quiz_state.get(chat_id, {}).get("active", False):
            await calculate_scores(chat_id, context)
        if chat_id in quiz_state:
            quiz_state.pop(chat_id, None)

      

    except Exception as e:
        print("Exception in sending poll:", e)

async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        answer = update.poll_answer
        user_id = str(answer.user.id)
        if answer.user.username:
            username = f"@{answer.user.username}" 
        elif answer.user.first_name:
            username = answer.user.first_name  
        else:
            username = str(user_id)
        selected_options = answer.option_ids
    

        for chat_id, chat_quiz in quiz_state.items():
            if "polls" not in chat_quiz:
                print(f"No 'polls' key in chat_quiz for chat_id {chat_id}")
                continue
            for poll in chat_quiz["polls"]:
                if poll["poll_id"] == answer.poll_id:
                    correct_answer = poll["correct_answer"]
                    options = poll["options"]

                    selected_answer = options[selected_options[0]]  # Assume single choice
                    if selected_answer == correct_answer:
                        # Initialize chat-specific score tracking
                        if chat_id not in quiz_scores:
                            quiz_scores[chat_id] = {}

                        if user_id not in quiz_scores[chat_id]:
                            quiz_scores[chat_id][user_id] = {"username": username, "score": 0}

                        quiz_scores[chat_id][user_id]["score"] += 1
                     

                    return  # Exit after finding the correct poll

    except Exception as e:
        print(f"Error in handle_poll_answer: {e}")

async def my_rank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the total score, rounds played, total groups, and rank of a user across all groups."""
    try:
        user_id = str(update.message.from_user.id)
        username = update.message.from_user.username or update.message.from_user.first_name or user_id
        username = escape_markdown(username)

        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found.")
            return

        user_scores = {}
        user_groups = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, chat_id, stored_user_id, stored_username, score, rounds = row
                stored_user_id = str(stored_user_id)

                if stored_user_id not in user_scores:
                    user_scores[stored_user_id] = {"score": 0, "rounds": 0}
                    user_groups[stored_user_id] = set()

                user_scores[stored_user_id]["score"] += int(score)
                user_scores[stored_user_id]["rounds"] += int(rounds)
                user_groups[stored_user_id].add(chat_id)


        sorted_users = sorted(user_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        rank = None
        for idx, (uid, data) in enumerate(sorted_users, start=1):
            if uid == user_id:
                rank = idx
                break

        if rank:
            total_score = user_scores[user_id]["score"]
            total_rounds = user_scores[user_id]["rounds"]
            total_groups = len(user_groups[user_id])
            totalperson = len(sorted_users) + 500
            response = (
                f"🏅 *Your Rank*\n\n"
                f"🏆🏆 {rank} out of {totalperson}\n"
                f"👤 *Username:* {username}\n"
                f"🎯 *Total Score:* {total_score}\n"
                f"🔄 *Rounds Played:* {total_rounds}\n"
                f"🧑‍🤝‍🧑 *Total Groups:* {total_groups}"
                
            )
        else:
            response = "You haven't participated in any quizzes yet."

        await update.message.reply_text(response, parse_mode="Markdown")

    except Exception as e:
        print(f"Error in my_rank: {e}")

async def topgrp_scorer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the top 10 users from the current group along with the total number of participants."""
    try:
        chat_id = str(update.message.chat_id)

        # Load Excel file
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found for this group.")
            return

        group_scores = {}
        total_users = set()

        # Collect scores for this specific group
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, stored_chat_id, user_id, username, score, _ = row
                if str(stored_chat_id) == chat_id:
                    total_users.add(user_id)  # Count unique users in this group
                    if user_id in group_scores:
                        group_scores[user_id]["score"] += int(score)
                    else:
                        group_scores[user_id] = {"username": escape_markdown(username), "score": int(score)}

        if not group_scores:
            await update.message.reply_text("No scores available for this group.")
            return

        # Sort and get top 10
        sorted_group_scores = sorted(group_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        leaderboard = f"🏆 *Top 10 Scorers in This Group* 🏆\n👥 *Total Members in Leaderboard:* `{len(total_users)}`\n\n"
        
        for rank, (user_id, data) in enumerate(sorted_group_scores[:10], start=1):
            leaderboard += f"{rank}\\) *{data['username']}* \\- `{data['score']} points`\n"

        await update.message.reply_text(leaderboard, parse_mode="MarkdownV2")

    except Exception as e:
        print(f"Error in top1grp_scorer: {e}")

def update_user_score(chat_id, correct_users,file):
    """
    Update user scores in an Excel file. If the user exists in the same chat_id, update their score.
    If the user exists in a different chat_id, create a new row.
    """
    try:
        game_round = 1
        try:
            workbook = load_workbook(file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Scores"
            sheet.append(["srno", "chatid", "Idnumber", "Username", "Score", "round"])
        existing_scores = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:  # Ensure row has all required columns
                sr_no, existing_chat_id, user_id, username, score, round = row
                existing_scores[(str(existing_chat_id), str(user_id))] = {
                    "sr_no": sr_no,
                    "chat_id": str(existing_chat_id),
                    "username": username,
                    "score": int(score),
                    "round": int(round)
                }
        for user_id, data in correct_users.items():
            username = data["username"]
            new_score = data["score"]
            key = (str(chat_id), str(user_id))
            if key in existing_scores:
                existing_scores[key]["username"] = username
                existing_scores[key]["score"] += new_score
                existing_scores[key]["round"] += game_round
            else:
                sr_no = len(existing_scores) + 1
                existing_scores[key] = {
                    "sr_no": sr_no,
                    "chat_id": str(chat_id),
                    "username": username,
                    "score": new_score,
                    "round": game_round
                }
        sheet.delete_rows(2, sheet.max_row)
        for (chat_id, user_id), data in existing_scores.items():
            sheet.append([data["sr_no"], chat_id, user_id, data["username"], data["score"], data["round"]])
        workbook.save(file)

    except Exception as e:
        print(f"Error updating scores: {e}")

async def calculate_scores(chat_id, context):
    global groupsendid
    if chat_id not in quiz_scores or not quiz_scores[chat_id]:
        quiz_state.pop(chat_id, None)
        quiz_scores.pop(chat_id, None)
        await context.bot.send_message(chat_id, "No one Selected the Correct Option in the quiz.")
        return

    update_user_score(chat_id, quiz_scores[chat_id],SCORE_FILE)
    update_user_score(chat_id, quiz_scores[chat_id],MNTH_SCORE_FILE)
    
    if os.path.exists(SCORE_FILE):
        with open(SCORE_FILE, 'rb') as file:
            await context.bot.send_document(chat_id=groupsendid, document=file)
    if os.path.exists(MNTH_SCORE_FILE):
        with open(MNTH_SCORE_FILE, 'rb') as file:
            await context.bot.send_document(chat_id=groupsendid, document=file)
    sorted_scores = sorted(quiz_scores[chat_id].items(), key=lambda x: x[1]["score"], reverse=True)

    leaderboard = f"🏆 *Quiz Results* 🏆\n\n"

    for rank, (user_id, data) in enumerate(sorted_scores, start=1):
        username = escape_markdown(data["username"])  # Removed version=2
        leaderboard += f"{rank}\\) *{username}* \\- `{data['score']} points`\n"
    
    leaderboard = leaderboard + "\n Start Quiz again with /startquiz"

    await context.bot.send_message(chat_id, leaderboard, parse_mode="MarkdownV2")

    quiz_state.pop(chat_id, None)
    quiz_scores.pop(chat_id, None)
  
def escape_markdown(text):
    """Escape special characters for Telegram MarkdownV2."""
    escape_chars = r"_*[]()~`>#+-=|{}.!\\"
    return re.sub(r"([{}])".format(re.escape(escape_chars)), r"\\\1", text)

async def month_topper(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the top 10 scorers across all groups, including total participants."""
    try:
        try:
            workbook = load_workbook(MNTH_SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found.")
            return

        global_scores = {}
        total_users = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, chat_id, user_id, username, score, _ = row
                total_users.add(user_id) 
                if user_id in global_scores:
                    global_scores[user_id]["score"] += int(score)
                else:
                    global_scores[user_id] = {"username": escape_markdown(username), "score": int(score)}

        if not global_scores:
            await update.message.reply_text("No scores available.")
            return
        sorted_global_scores = sorted(global_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        leaderboard = f"🌍 *This \\ Month Top Scorer Of Overall* 🌍\n👥 *Total Participants:* `{len(total_users)}`\n\n"
        
        for rank, (user_id, data) in enumerate(sorted_global_scores[:10], start=1):
            leaderboard += f"{rank}\\) *{data['username']}* \\- `{data['score']} points`\n"

        await update.message.reply_text(leaderboard, parse_mode="MarkdownV2")

    except Exception as e:
        print(f"Error in all_time_topper: {e}")

async def all_time_topper(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Returns the top 10 scorers across all groups, including total participants."""
    try:
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            await update.message.reply_text("No scores found.")
            return

        global_scores = {}
        total_users = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                _, chat_id, user_id, username, score, _ = row
                total_users.add(user_id) 
                if user_id in global_scores:
                    global_scores[user_id]["score"] += int(score)
                else:
                    global_scores[user_id] = {"username": escape_markdown(username), "score": int(score)}

        if not global_scores:
            await update.message.reply_text("No scores available.")
            return
        sorted_global_scores = sorted(global_scores.items(), key=lambda x: x[1]["score"], reverse=True)
        leaderboard = f"🌍 *All\\-Time Top 10 Scorers* 🌍\n👥 *Total Participants:* `{len(total_users)}`\n\n"
        
        for rank, (user_id, data) in enumerate(sorted_global_scores[:10], start=1):
            leaderboard += f"{rank}\\) *{data['username']}* \\- `{data['score']} points`\n"

        await update.message.reply_text(leaderboard, parse_mode="MarkdownV2")

    except Exception as e:
        print(f"Error in all_time_topper: {e}")

async def broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global registered_groups

    chat_id = update.effective_chat.id
    main_group_id = str(botManagementGroupId)
    to_remove = set()

    if chat_id != botManagementGroupId:
        return

    if not context.args:
        await update.message.reply_text("❗ Please provide a message.\nUsage:\n/broadcastmessage Hello! Click below to start.")
        return

    message = " ".join(context.args)

    # Replace with your actual bot username
    start_url = "https://t.me/slizzyy_bot?start=start"
    keyboard = [[InlineKeyboardButton("Send Now", url=start_url)]]

    async def send_to_group(gid):
        gid_str = str(gid)
        try:
            previous_id = group_data.get(gid_str, {}).get("last_message_id")
            if previous_id:
                try:
                    await context.bot.delete_message(chat_id=gid, message_id=previous_id)
                except Exception as e:
                    print(f"Failed to delete old message in {gid}: {e}")

            sent_msg = await context.bot.send_message(
                chat_id=gid,
                text=message,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            group_data.setdefault(gid_str, {})["last_message_id"] = sent_msg.message_id
            return True, gid
        except Exception as e:
            print(f"Error sending to group {gid}: {e}")
            if "bot was kicked" in str(e) or "chat not found" in str(e).lower():
                return False, gid
            return None, gid

    tasks = [asyncio.create_task(send_to_group(gid)) for gid in registered_groups]
    results = await asyncio.gather(*tasks)

    count = 0
    for success, gid in results:
        if success is True:
            count += 1
        elif success is False:
            to_remove.add(gid)

    if to_remove:
        registered_groups -= to_remove
        save_groups()
    save_group_data()

    await update.message.reply_text(f"✅ Custom message sent to {count} groups.")

async def send_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    global registered_groups
    if chat_id != botManagementGroupId:
        return
    main_group_id = str(botManagementGroupId)
    to_remove = set()

    if main_group_id not in group_data:
        await update.message.reply_text("\u274c No message/link set in main group.")
        return

    message = group_data[main_group_id].get("message", "")
    keyboard = [[InlineKeyboardButton("Start Quiz", callback_data='New_1')]]

    async def send_to_group(gid):
        gid_str = str(gid)
        try:
            previous_id = group_data.get(gid_str, {}).get("last_message_id")
            if previous_id:
                try:
                    await context.bot.delete_message(chat_id=gid, message_id=previous_id)
                except Exception as e:
                    print(f"Failed to delete old message in {gid}: {e}")
            sent_msg = await context.bot.send_message(
                chat_id=gid,
                text=message,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            group_data.setdefault(gid_str, {})["last_message_id"] = sent_msg.message_id
            return True, gid
        except Exception as e:
            print(e)
            if "bot was kicked" in str(e) or "chat not found" in str(e).lower():
                return False, gid
            return None, gid  # some other error

    # Schedule all tasks
    tasks = [asyncio.create_task(send_to_group(gid)) for gid in registered_groups]
    results = await asyncio.gather(*tasks)

    count = 0
    for success, gid in results:
        if success is True:
            count += 1
        elif success is False:
            to_remove.add(gid)

    if to_remove:
        registered_groups -= to_remove
        save_groups()
    save_group_data()

    await update.message.reply_text(f"\u2705 Message sent to {count} groups.")

async def register_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    if chat.type in ["group", "supergroup"]:
        if chat.id not in registered_groups:
            registered_groups.add(chat.id)
            save_groups()
            if os.path.exists(GROUPS_FILE):
                with open(GROUPS_FILE, 'rb') as file:
                    await context.bot.send_document(chat_id=groupsendid, document=file)
            
async def add_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id != botManagementGroupId:
        return
    text = " ".join(context.args)
    if not text:
        await update.message.reply_text("Usage: /addmessage <your message>")
        return

    group_data.setdefault(str(chat_id), {})["message"] = text
    save_group_data()
    await update.message.reply_text("\u2705 Message set for the main group.")

async def add_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global new_timer
    chat_id = update.effective_chat.id
    if chat_id != botManagementGroupId:
        return
    input_value = " ".join(context.args)
    try:
        float(input_value)  # can also use int(input_value) if you want only integers
        new_timer = input_value
        await update.message.reply_text(f"New timer set to {new_timer}.")
    except ValueError:
        await update.message.reply_text("Invalid input. Please enter a numeric value for the timer.")


async def show_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id != botManagementGroupId:
        return
    data = group_data.get(str(chat_id), {})
    message = data.get("message", "No message set.")
    link = data.get("link", "")

    keyboard = []
    keyboard = [[InlineKeyboardButton("Start Quiz",callback_data='New_1')]]
    await update.message.reply_text(
        message,
        reply_markup=InlineKeyboardMarkup(keyboard) if keyboard else None
    )

async def add_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global newuploadedexcelfile
    chat_id = update.effective_chat.id
    if chat_id != botManagementGroupId:
        return
    file_name = context.args[0]

    if file_name not in ALLOWED_FILES:
        await update.message.reply_text("❌ Invalid file name. Please choose from:\n" + "\n".join(sorted(ALLOWED_FILES)))
        return
    
    newuploadedexcelfile = " ".join(context.args)
    newuploadedexcelfile = "Data/"+newuploadedexcelfile + ".xlsx"
    print(newuploadedexcelfile)

    group_data.setdefault(str(chat_id), {})["link"] = newuploadedexcelfile
    save_group_data()
    await update.message.reply_text("\u2705 Link set for the main group.")


async def add_promo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global newuploadedexcelfile,Promotion
    chat_id = update.effective_chat.id
    if chat_id != botManagementGroupId:
        return
    msg = context.args[0]
    if msg.lower() =="yes":
        Promotion = True
        await update.message.reply_text("✅ Promotion enabled. You can now send promotional messages in the quiz.")
        return

    elif msg.lower() == "no":
        Promotion = False
        await update.message.reply_text("❌ Promotion disabled. No promotional messages will be sent in the quiz.")
        return

async def countdown_and_close_poll(chat_id, poll_message, context):
    try:
        selectedtime = quiz_state[chat_id]["selectedtime"]
        if not isinstance(selectedtime, int):
            try:
                selectedtime = int(selectedtime)
            except:
                selectedtime = 15

        if selectedtime < 10:
            selectedtime = 15
        await asyncio.sleep(selectedtime)
        closed_poll = await poll_message.stop_poll()
        

        for poll in quiz_state[chat_id]["polls"]:
            if poll["poll_id"] == poll_message.poll.id:
                meaning = poll["meaning"]
                if str(meaning).strip().lower() == "nan" or not str(meaning).strip():
                    break
                if meaning:
                    await context.bot.send_message(chat_id, text=f"Meaning: {meaning}")
                break
        total_votes = sum(option.voter_count for option in closed_poll.options)
        if total_votes == 0:
            quiz_state[chat_id]["consecutive_unanswered"] += 1
        else:
            quiz_state[chat_id]["consecutive_unanswered"] = 0
        if quiz_state[chat_id]["consecutive_unanswered"] >= 3:
            await context.bot.send_message(chat_id, text="❌ Quiz canceled due to inactivity. Restart with /startquiz")
            quiz_state[chat_id]["active"] = False
            await calculate_scores(chat_id, context) 
            return 
    except Exception as e:
        print(f"Error in countdown_and_close_poll: {e}")

async def handle_jsonFile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global registered_groups
    chat_id = update.effective_chat.id

    if chat_id == groupsendid and update.message.reply_to_message and update.message.reply_to_message.document:
        try:
            file = await context.bot.get_file(update.message.reply_to_message.document.file_id)
            downloaded_path = await file.download_to_drive()

            # Replace existing userscore.xlsx
            if os.path.exists(GROUPS_FILE):
                os.remove(GROUPS_FILE)

            os.rename(downloaded_path, GROUPS_FILE)
            
            if os.path.exists(GROUPS_FILE):
                with open(GROUPS_FILE, "r") as f:
                    registered_groups = set(json.load(f))
            await update.message.reply_text("✅ Score file updated successfully.")
        except Exception as e:
            await update.message.reply_text(f"❌ Failed to update score file:\n`{e}`", parse_mode="Markdown")
    else:
        await update.message.reply_text("⚠️ Please reply to an .xlsx file in the allowed group.")

async def cancel_quiz_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global quiz_tasks
    chat_id = update.message.chat_id
    if chat_id in quiz_state and quiz_state[chat_id]["active"]:
        quiz_state[chat_id]["active"] = False
        quiz_scores.pop(chat_id, None)
        quiz_state.pop(chat_id, None)
        if chat_id in quiz_tasks:
            task = quiz_tasks[chat_id]
            if not task.done():
                task.cancel()
        await context.bot.send_message(chat_id, text="The quiz has been canceled. You can restart with /startquiz.")
    else:
        if chat_id in quiz_state:
            quiz_state.pop(chat_id, None)
            quiz_scores.pop(chat_id, None)
        await context.bot.send_message(chat_id, text="No active quiz to cancel.")

def stringify(value):
    if value is None or value == "":
        return "None"
    return str(value)
